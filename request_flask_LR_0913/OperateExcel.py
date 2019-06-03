# encoding=utf-8
import collections
import os

from openpyxl import Workbook,load_workbook
from flask import json
import pythoncom
import win32com.client



excelFile = os.getcwd()+"\\test.xlsx"
imageRootPath = os.getcwd()+"\\upload"

def getRowAndFile(contentId):
    fileName = ""
    row = 0
    workBook = load_workbook(excelFile)
    sheetObj = workBook.get_sheet_by_name(workBook.sheetnames[0])
    maxRow = sheetObj.max_row
    if maxRow >1:
        for i in range(2,maxRow+1):
            if int(contentId) == int(sheetObj.cell(row=i,column=1).value):
                fileName = sheetObj.cell(row=i,column=3).value
                row = i
                break
    workBook.close()
    return row,fileName

# 创建文件
def createExcel():
    if not os.path.exists(excelFile):
        workBook =Workbook()
        sheet = workBook.active
        sheet.cell(row=1,column=1).value = "id"
        sheet.cell(row=1,column=2).value = "token"
        sheet.cell(row=1,column=3).value = "url"
        workBook.save(excelFile)

#         获取最大的id
def getMaxRowValue():
    # 内容的contentID是从1开始的，表格的1是表头
    workBook = load_workbook(excelFile)
    sheerObj = workBook.get_sheet_by_name(workBook.sheetnames[0])
    maxRow = sheerObj.max_row
    maxRowValue = 0
    if maxRow > 1:
        maxRowValue = sheerObj.cell(row = maxRow,column = 1).value
    workBook.close()
    return maxRowValue




# 添加内容
def addContent(contentValue,file):
    # 判断文件是否存在，不存在需要创建一个新的excel文件
    createExcel()
#     获取最大的contentID
    conrentId = getMaxRowValue() + 1
    # 保存file
    # imagePath = imageRootPath +"\\"+str(conrentId)+".png"
    # file.save(imagePath)
    # file.close()
    workBook = load_workbook(excelFile)
    sheerObj = workBook.get_sheet_by_name(workBook.sheetnames[0])
    sheerObj.append((conrentId,contentValue))
    workBook.save(excelFile)
    workBook.close()
    return {"message":"add content success","Id":conrentId,"status":1,"token":int(contentValue)}


# 查询内容
def queryContent(contentId = None):
    resultList = []

    workBook = load_workbook(excelFile)
    sheerObj = workBook.get_sheet_by_name(workBook.sheetnames[0])
    maxRow = sheerObj.max_row
    if maxRow>1:
        for i in range(2,maxRow+1):
            contentInfo = collections.OrderedDict()
            contentInfo["contentId"] = sheerObj.cell(row=i,column=1).value
            contentInfo["contentValue"] = sheerObj.cell(row=i,column=2).value
            contentInfo["contentImageFile"] = sheerObj.cell(row=i,column=3).value
            if contentId !=None:
                if int(contentId) == int(sheerObj.cell(row=i,column=1).value):
                    resultList.append(contentInfo)
            else:
               resultList.append(contentInfo)
        return json.dumps(resultList,ensure_ascii = False,indent = 4)
    else:
        return resultList


# 删除内容
def deleteContent(contentId):
    #打开Excel
    pythoncom.CoInitialize()
    workBook = win32com.client.Dispatch("Excel.Application").Workbooks.Open(excelFile)
    try:
        sheetObj = workBook.Worksheets("Sheet")
        row,fileName = getRowAndFile(contentId) #获取删除的行和文件
        if row > 1:
            #先删除数据
            sheetObj.Rows(row).Delete()
            #删除文件
            os.remove(imageRootPath+"\\"+fileName)
            return {"message":"delete success","status":0,"id":str(contentId)}
        else:
            return {"message":"delete success","status":1,"id":str(contentId)}
        return {"message":"delete failed","status":0}
    except Exception,e:
        print str(e)
    finally:
        workBook.Close(SaveChanges=1)





















