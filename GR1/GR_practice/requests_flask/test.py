# encoding=utf-8
import os
from openpyxl import load_workbook,Workbook

print os.path.exists("1.jpg")

# sheet = out.create_sheet("test0827",index =1)
workbook = load_workbook("test0827.xlsx")
for i in workbook.get_sheet_names():
    print i

sheet1 = workbook.get_sheet_by_name("Sheet1")
# sheet1.cell(row=3,column=5).value = 'sfdf'

allrow = sheet1.max_row
for i in xrange(allrow):
        print sheet1.cell(row=i+1,column=1).value
print sheet1.max_row
workbook.save("test0827.xlsx")
a = "a;b;c;d"
print a.split(';')