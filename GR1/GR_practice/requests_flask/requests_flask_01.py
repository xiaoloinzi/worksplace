# encoding=utf-8
from flask import Flask,request,jsonify,send_file,Response,abort,make_response
import os
import random
from openpyxl import load_workbook
# 【20170806课后练习】 模拟微信朋友圈的服务端，需要满足以下场景：
# 1.发表内容时带上图片信息，内容包括（具体内容和id，id指定全局唯一标识），客户端请求后服务端将信息保存在excel中
# 2.删除无效的内容
# 3.对朋友圈进行查询操作，返回所有的查询结果（内容在控制台进行打印，文件保存在临时目录）
# 【注意】
# 客户端：使用时建议做一个流程，即发表->查询->删除，这样来测试服务端--Publish -> Query -> Delete
# 测试异常：如果进行查询或者删除不存在的资源时，提示资源不存在并且返回错误码为404

app = Flask(__name__)


@app.route("/publish",methods=["post"])
def publish():
    filename = request.files["fileName"]
    data = request.form.items()
    workbook = load_workbook("test0827.xlsx")
    sheet1 = workbook.get_sheet_by_name("Sheet1")
    file = "1.png"
    token = random.randint(10000,200000)
    # 确保文件不被覆盖
    while os.path.exists(file):
        file = str(random.randint(1,100)) + str(file)
    filename.save(os.getcwd()+"\\"+file)
    # 获得最大的行数
    rowa = sheet1.max_row
    print rowa
    for i in xrange(rowa):
        if sheet1.cell(row=i+1,column=1).value == data[0][1]:
            oldstr = sheet1.cell(row=i+1,column=2).value
            oldfile = sheet1.cell(row=i+1,column=3).value
            sheet1.cell(row=i+1,column=2).value = oldstr+";"+data[1][1]
            sheet1.cell(row=i+1,column=3).value = oldfile +";"+file
            break
        elif i+1 == rowa:
            sheet1.cell(row=rowa+1,column=1).value = data[0][1]
            sheet1.cell(row=rowa+1,column=2).value = data[1][1]
            sheet1.cell(row=rowa+1,column=3).value = file
            break
    workbook.save("test0827.xlsx")
    return str({"message":"Published successfully","data":";".join(['='.join(item) for item in data]),"filename":file,"token":token})

@app.errorhandler(404)
def queryerror(message):
    userid = request.args.get("userID")
    result = make_response(jsonify({"message":"Sorry, your account does not exist or The deleted data does not exist","userid":userid}))
    result.status_code = 404
    result.headers['X-Parachutes'] = None
    return result


@app.route("/query",methods=["get"])
def query():
    # 获取文件名
    filename = request.args.get("fileName")
    # 获取id
    userId = request.args.get("userID")
    # 读取excel表
    workbook = load_workbook("test0827.xlsx")
    sheet1 = workbook.get_sheet_by_name("Sheet1")
    rowa = sheet1.max_row
    for i in xrange(rowa+1):
        if sheet1.cell(row=i+1,column=1).value == userId:
            strnum = sheet1.cell(row=i+1,column=2).value
            filenum = sheet1.cell(row=i+1,column=3).value
            break
        elif i == rowa:
            abort(404)
    fileList = filenum.split(';')
    for files in xrange(len(fileList)):
        if str(fileList[files]) == filename:
            print str(fileList[files])
            file = send_file(filename)
            mr =  make_response(file)
            print mr
            break
        elif files+1 == len(fileList):
            mr = make_response(jsonify({"message":"The file you need does not exist","filename":filename}))
    workbook.close()
    mr.headers['X-Parachutes'] = strnum
    return mr

@app.route("/delete",methods=["get"])
def delete():
    # 获取文件名
    filename = request.args.get("fileName")
    # 获取id
    userId = request.args.get("userID")

    usernum = request.args.get("stringNum")
    # 读取excel表
    workbook = load_workbook("test0827.xlsx")
    sheet1 = workbook.get_sheet_by_name("Sheet1")
    rowa = sheet1.max_row
    for i in xrange(rowa+1):
        if sheet1.cell(row=i+1,column=1).value == userId:
            strnum = sheet1.cell(row=i+1,column=2).value
            filenum = sheet1.cell(row=i+1,column=3).value
            fileList = filenum.split(';')
            strList = strnum.split(';')
            if filename in fileList and usernum in strList :
                fileList.remove(filename)
                strList.remove(usernum)
                sheet1.cell(row=i+1,column=3).value =  ";".join(fileList)
                sheet1.cell(row=i+1,column=2).value = ";".join(strList)
                mr = make_response(jsonify({"message":"delete success "}))
            else:
                abort(404)
            break
        elif i == rowa:
            abort(404)


    workbook.close()
    return mr

if __name__=="__main__":
    app.run(host="0.0.0.0",port=8888,debug=True)











