# encoding=utf-8
import os
import time
import datetime
import xlrd
import xlwt
import openpyxl
from xlutils.copy import copy
from openpyxl import load_workbook
from openpyxl.styles import Border,Side,Font,\
    PatternFill,Alignment,Protection
from openpyxl.styles import colors
from openpyxl.styles import Font,Color
from openpyxl.styles import colors
from openpyxl import Workbook
import sys
reload(sys)
sys.setdefaultencoding('utf-8')
# data = xlrd.open_workbook("D:\\tmp\\test0827.txt.xlsx")
# print type(data)
# tablelist = data.sheets()
# for i in tablelist:
#     print i
# print type(tablelist)
# tablename = data.sheet_by_name(u'员工信息表')
# print tablename
# tableindex = data.sheet_by_index(0)
# print u'索引号为0的工作表为：',tableindex.name
# for i in tablename.row_values(0):
#     print i,
# print
# for i in xrange(1,3):
#     for j in tableindex.row_values(i):
#         print j,
#     print
# tablerow = tableindex.row_values(1)
# print type(tablerow),len(tablerow)
#
# tablecol = tableindex.col_values(0)
# print tablecol,type(tablecol)
# for i in tablecol:
#     print i
# print len(tablecol)
# print u'行数：',tableindex.nrows
# print u'列数：',tableindex.ncols
# print u'获取单元格中的值：'
# cell = tableindex.cell(0,0)
# print cell
# print cell.value

# workbook = xlwt.Workbook(encoding = 'utf-8')
# print type(workbook)
#
# worksheet = workbook.add_sheet('tet.xlsx')
# print u'创建Excel表的表名是：',worksheet.name
#
# style = xlwt.easyxf('pattern:pattern solid,fore_color green;')
#
# worksheet.write(r=1,c=2,label=u'这是员工测试',style=style)
# workbook.save('D:\\tmp\\Excel_test.xls')

# workbook1 = xlwt.Workbook(encoding='utf-8')
# worksheet1 = workbook1.add_sheet('uu')
# worksheet1.write(r=3,c=5,label=u'测试')
# workbook1.save('D:\\tmp\\笑笑.xls')

# workread = xlrd.open_workbook('D:\\tmp\\test0827.txt.xlsx')
# print type(workread.sheets())
# worklist = workread.sheets()
# for i in worklist:
#     print i
# workname = workread.sheet_by_name('员工信息表')
# print type(workname)
# workindex = workread.sheet_by_index(0)
# print workindex
# workrow = workread.sheet_by_index(0).row_values(2)
# for i in workrow:
#     print i,
# print
# workcol = workread.sheet_by_index(0).col_values(2)
# for i in workcol:
#     print i,
# print
# print '行数：',workread.sheet_by_index(0).nrows
# print '列数：',workread.sheet_by_index(0).ncols
# rows1 =  workread.sheet_by_index(0).row(1)
# for i in rows1:
#     print i,
# print
# print workread.sheet_by_index(0).cell(1,2).value

# workwrite = xlwt.Workbook(encoding='utf-8')
# print type(workwrite)
# worksheet = xlwt.Workbook(encoding='utf-8').add_sheet('test1.xlsx')
# print u'创建的表名：',worksheet.name
# worksheet.write(r=1,c=2,label='test0827.txt')
# workwrite.save('D:\\tmp\\ec_te.xls')

# readbook = xlrd.open_workbook('D:\\tmp\\test0827.txt.xlsx')
# copybook = copy(readbook)
# for i in range(len(readbook.sheets())):
#     sheet = copybook.get_sheet(i)
#     readsheet = readbook.sheet_by_index(i)
#     for row in range(readsheet.nrows):
#         for col in range(readsheet.ncols):
#             cell = readsheet.cell(row,col).value
#             print 'type is ',type(cell),cell
#             if type(cell) in (str,unicode):
#                 res = cell.upper()
#                 sheet.write(row,col,res)
# copybook.save("D:\\tmp\\t2.xls")

# workbook = load_workbook('D:\\tmp\\test0827.txt.xlsx')
# print workbook
# bookNameList = workbook.get_sheet_names()
# print type(bookNameList)
# for i in bookNameList:
#     print i
# sheet1 = workbook.get_sheet_by_name(u'员工信息表')
# print type(sheet1)

# workbook = load_workbook('D:\\tmp\\test0827.txt.xlsx')
# bookNameList = workbook.get_sheet_names()
# sheet1 = workbook.get_sheet_by_name(u'员工信息表')
# all_row = sheet1.rows
# all_col = sheet1.columns
# # print all_row
# # print type(all_row)#,len(all_row)
# # for i in all_row:
# #     print '\n',type(i)
# #     for j in i:
# #         print j,j.value
# print type(all_col)
# for i in all_col:
#     print type(i)
#     for j in i:
#         print type(j),j,j.value

# workbook = load_workbook('D:\\tmp\\test0827.txt.xlsx')
# bookNamelist = workbook.get_sheet_names
# sheet1 = workbook.get_sheet_by_name(u'员工信息表')
# print u'***************通过行获取单元格**************'
# all_row = sheet1.rows
# for i in all_row:
#     for j in xrange(len(i)):
#         if i[j].coordinate == 'C3':
#             print u'单元格C3的值：',i[j].value
#             break
#
# print u'***************通过列获取单元格**************'
# all_col = sheet1.columns
# for i in all_col:
#     for j in xrange(len(i)):
#         if i[j].coordinate == 'A2':
#             print u'单元格中A2的值：',i[j].value
#             break
# workbook = load_workbook('D:\\tmp\\test0827.txt.xlsx')
# bookNameList = workbook.get_sheet_names()
# sheet1 = workbook.get_sheet_by_name(u'员工信息表')
# b3 = sheet1.cell(coordinate='B3')
# print u'单元格B3的值：',b3.value
# # c2 = sheet1.cell(row=2,column=3)
# # print u'单元格C2的值：',c2.value
# # cell = sheet1.cell(coordinate='C3',row=3,column=3)
# # print cell.value
# sheet1.cell(coordinate='B3').value = 300
# print u'单元格B3的值：',b3.value
# workbook.save('D:\\tmp\\test0827.txt.xlsx')
# rows = sheet1.rows
# # print rows[:2]
# area_sheet = sheet1['A1':'C4']
# print type(area_sheet)
# for i in area_sheet:
#     print i,i[0].value
# out = openpyxl.Workbook()
# sheet2 = out.create_sheet(u'员工工资表',index=1)
# print u'创建表的表名：',sheet2.title
# print type(sheet2)
# print sheet2.cell(row=1,column=2).value
# sheet2.cell(row = 1,column=2).value=1
# print sheet2.cell(row=1,column=2).value
# print sheet2.cell(coordinate='C4').value
# sheet2.cell(coordinate='C4').value=356
# print sheet2.cell(coordinate='C4').value
# out.save('D:\\glroyroadtest.xlsx')

# sheet2["A4"].value = 2
# sheet2["B4"].value = 'test0827.txt'
# sheet2["C4"].value = 'this is a test0827.txt'
# sheet2['D4'].value = u'这是员工测试'
# now = time.strftime('%Y-%m-%d %H:%M:%S',time.localtime())
# sheet2['C2'] = now
# for i in xrange(1,5):
#     n = 'D'+str(i)
#     sheet2[n]= str(i)+'kk'
# sheet2.append(('a',1,2,3,u'中国'))
# out.save("D:\\tmp\\glroyroadtest.xlsx")

# font = Font(size=15,
#             bold=True,
#             italic=False,
#             vertAlign=None,
#             underline='none',
#             strike=False,
#             color='FF008B00'
# )
# fill = PatternFill(
#     fill_type=None,
#     start_color='FFFF3030',
#     end_color='FF000000'
# )
# alignment = Alignment(
#     horizontal='general',
#     vertical='bottom',
#     text_rotation=0,
#     wrap_text=False,
#     shrink_to_fit=False,
#     indent = 0
# )
# border = Border(left=Side(),
#                 right=Side(),
#                 top=Side(),
#                 bottom=Side())
# protection = Protection(locked=True,
#                         hidden=False)
# out = openpyxl.Workbook()
# sheet2 = out.create_sheet(u'员工工资表',index=0)
# print u'创建的表的表名：',sheet2.title
# sheet2.append(('a',1,2,3,u'中国'))
# sheet2['A3'] = u'test0827.txt'
# sheet2['A3'].font = font
# sheet2['A5'].border = border
# sheet2['A3'].alignment = alignment
# # print sheet2['e3'].border
# out.save('D:\\tmp\\we.xlsx')

wb = Workbook()
ws = wb.active

ws.append([1,2,3])
a1 = ws['A1']
a2 = ws['B1']
a3 = ws['C1']
d4 = ws['D4']
ft = Font(color=colors.RED)
a1.font = ft
a2.font = ft
a3.font = ft
d4.font = ft
a1.font = Font(color=colors.RED,italic=True)
ws['A1'] = 42
wb.save('D:\\tmp\\w1.xlsx')










