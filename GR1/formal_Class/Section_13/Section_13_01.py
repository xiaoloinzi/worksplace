# encoding=utf-8
import time
import xlrd
import xlwt
import openpyxl
import xlutils
from openpyxl.styles import PatternFill,Alignment,Font,colors

# 2003之前：Excel：xls
# 2003之后：Excel：xlsx
# xlrd:读取的模块：xls,xlsx
# xlwt:写分模块：xls
# openpyxl:既能读也能写，只能操作xlsx

wordBook = xlrd.open_workbook('D:\\tmp\\t2.xls')#整个Excel
print type(wordBook)
sheetList = wordBook.sheets()
print sheetList
sheet1 = sheetList[0]#Excel的sheet
print sheet1
sheet2 = wordBook.sheet_by_index(0)
print sheet2
sheet3 = wordBook.sheet_by_name('test0827.txt')
print sheet3

rowList = sheet3.row_values(9)
# 通过下标取得某个单元格的值
print rowList#整行的值
colList = sheet3.col_values(1)
print colList
cell = sheet3.cell(1,4)#定位
print cell.value#整列的值
#workbook->整个excle
# sheet->excle 的sheet
# rowList ->整行的值
# cell->某一个单元格的值
#
# 1、获取第1个sheet 中B10的值，用不少于两种方法。
print colList[9]
print rowList[1]
print sheet3.cell(9,1).value
print u'行数：',sheet3.nrows
print u'列数：',sheet3.ncols

# 2、读取‘A2：D9’区域的所有数据
for i in xrange(2,10):
    cell1 = sheet3.row_values(i)
    for j in cell1[0:4]:
        print j,
    print

for  i in xrange(1,9):
    for j in xrange(4):
        c2 = sheet3.cell(i,j)
        print c2.value
    print

# xlwt
# 1、生成workbook
# 2、生成sheet
# 3、生成eExcel
#
workBook = xlwt.Workbook(encoding='utf-8')
print type(workBook)
sheet1 = workBook.add_sheet('test3')
sheet2 = workBook.add_sheet('test2')
sheet4 = workBook.add_sheet('test4')
print 'the sheet name is ',sheet1.name
sheet1.write(r =2,c=2,label='test0827.txt data')
sheet1.write(r =1,c=2,label=u'测试')
sheet1.write(r =3,c=2,label=u'光荣之路')
for i in xrange(4):
    for j in xrange(5):
        str1 = u'测试'+str(i)
        sheet2.write(r=i,c=j,label=str1)

# workBook.save('D:\\tmp\\t3.xls')
# # 1、在上一题中，把‘A2：D9’区域读取的内容，写入一个xls 的表格中，
# # 新的excle表格写入的区域是'A1：D8'
for  i in xrange(1,9):
    for j in xrange(4):
        c2 = sheet3.cell(i,j).value
        sheet1.write(r=(i-1),c=j,label=c2)
for  i in xrange(1,9):
    for j in xrange(4):
        c2 = sheet3.cell(i,j).value
        if isinstance(c2,(str,unicode)):
            c2 = c2.lower()
        sheet4.write(r=(i-1),c=j,label=c2)
#
#
#
#
# # 2、复制excle1 的sheet1 到excles2 的sheet1
#
a = sheet3.nrows
b = sheet3.ncols

for  i in xrange(a):
    for j in xrange(b):
        c2 = sheet3.cell(i,j).value
        sheet4.write(r=i,c=j,label=c2)

workBook.save('D:\\tmp\\t3.xls')


openpyxl
# workbook-sheet-cell
workbook = openpyxl.load_workbook('D:\\tmp\\test0827.txt.xlsx')
workbook1 = openpyxl.load_workbook(u'D:\\tmp\\测试.xlsx'.encode('gbk'))
sheet1test = workbook1.get_sheet_by_name(u'员工信息表')
rowData1 = sheet1test.rows
for i in rowData1:
    print i
print type(workbook)
sheetList = workbook.get_sheet_names()
for i in sheetList:
    print i
sheet1 = workbook.get_sheet_by_name('test0827.txt')
rowData = sheet1.rows
# 行从1开始，列也是
for i in rowData:
    # print type(i)
    # print i
    for j in i:
        print type(j)
        print j.coordinate,
        print j.value,
    print

colData1 = sheet1test.columns
for i in colData1:
    # print type(i)
    # print i
    for j in i:
        # print type(j)
        print j.coordinate,
        print j.value,
    print

# 1、创建一个xlsx，通过rows方法，找出D8单元格的数据

rowData = sheet1.rows
for i in rowData:
    # print type(i)
    # print i
    for j in i:
        # print type(j)
        if j.coordinate=='D8':
            print j.value

# 习题，通过列找D8

colData = sheet1.columns
for i in colData:
    # print type(i)
    # print i
    for j in i:
        # print type(j)
        if j.coordinate=='D8':
            print j.value

# 更新一个单元格的值
sheet1['A3'].value = 12
workbook.save('D:\\tmp\\test0827.txt.xlsx')

# C1 = sheet1.cell(coordinate=,row=,columns=)
c1 = sheet1.cell(coordinate='A7')
print c1.value
print c1.coordinate
c1 = sheet1.cell(row=1,columns=7)
print c1.value
print c1.coordinate

# 2、把C2~C7的单元格，value写coordinate的值

for i in xrange(2,8):
    c1 = sheet1.cell(row=3,column=i)
    print c1.value
    c1.value =c1.coordinate
    print c1.value

# 读一个区域的时候要注意有多行多列
str1 = sheet1['A2':'G7']
for i in str1:
    for j in xrange(len(i)):
        print i[j].value,
    print
    print i[0].value
    i[0].value = i[0].coordinate
    print i[0].value

book1 = openpyxl.Workbook()
book1sheet1 = book1.create_sheet('sheet1',index=0)
print type(book1sheet1)
print book1sheet1.title#sheet的名称

book1sheet1['A2'].value = u'测试写入成功'
print book1sheet1['A2'].value


str5 = time.strftime('%Y-%m-%d %H:%M:%S',time.localtime())
str4 = [[u'姓名',u'时间',u'分数'],['lily',time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),34]]
str2 = book1sheet1['A1':'C2']
s = 0
for i in str2:
    for j in xrange(len(i)):
        i[j].value= str4[s][j]
    s +=1

# 可以整行写入，但是不能保证在哪一行
book1sheet1.append([u'姓名',u'时间',u'分数'])
book1sheet1.append(['lily',time.strftime('%Y-%m-%d %H:%M:%S',time.localtime()),34])
#


# book1.save('D:\\tmp\\t10.xlsx')
# 合并单元格、单元格居中，修改背景颜色、字体的颜色
# 合并sheet.merge_cells(range_string="合并的区域"，start_row=开始的行
# start_column= 开始的列，end_row=结束的行，end_column = 结束的列）
book2 = openpyxl.Workbook()
book1sheet2 = book2.create_sheet('test3',index=0)
book1sheet2.merge_cells(range_string="A1:I1")

# 背景颜色
# PatternFill(patternType='',fgColor='',bgColor='',
#             fill_type='',
#             start_color='',
#             end_color='')
#
fill = PatternFill(patternType='solid',fgColor=colors.GREEN)
book1sheet2['A1'].fill = fill

# 字体的设置
# Font(name='',
#      sz='',
#      b='',
#      i='',
#      charset='',
#     )
book1sheet2['A1'].value = u'九九乘法表'
font = Font(color = colors.WHITE,size=14)
book1sheet2['A1'].font = font

# 对齐的方式
a1 = Alignment(horizontal='center')
book1sheet2['A1'].alignment = a1


# 写一个九九乘法表到excel中
for i in xrange(1,10):
    list1 = []
    for j in xrange(1,i+1):
        if i >= j:
            ste = '%d*%d'%(j,i)+'='+str(j*i)
            book1sheet2.cell(row=i+1,column=j).value = ste
    #     list1.append(ste)
    # book1sheet2.append(list1)
book2.save('D:\\tmp\\t11.xlsx')
