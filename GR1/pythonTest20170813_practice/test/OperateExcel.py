# encoding=utf-8
from openpyxl import Workbook,load_workbook
import os,collections,json
import win32com.client #excel,word
import pythoncom

excelFile = os.getcwd()+"\\test0827.xlsx"
imageRootPath = os.getcwd()+"\\upload"
#创建文件
def createExcel():
    if os.path.exists(excelFile) == False:
        workBook = Workbook()
        sheet = workBook.active
        sheet.cell(row=1,column=1).value = "编号"
        sheet.cell(row=1,column=2).value = "内容"
        sheet.cell(row=1,column=3).value = "图片路径"
        workBook.save(excelFile)

def getMaxRowValue():#内容的contentID是从1开始，表格中1是表头
    workBook = load_workbook(excelFile)
    sheetObj = workBook.get_sheet_by_name(workBook.sheetnames[0])
    maxRow = sheetObj.max_row
    maxRowValue = 0
    if maxRow > 1:
        maxRowValue = sheetObj.cell(row=maxRow,column=1).value
    workBook.close()
    return  maxRowValue
#添加内容
def addContent(contentValue,file):
    #判断文件是否存在，不存在需要创建一个新的excel文件
    createExcel()
    #每次获取最大行的第一列的值
    contentId = getMaxRowValue()+1
    #保存file
    imagePath =imageRootPath+"\\"+str(contentId)+".png"
    file.save(imagePath)
    file.close()
    workBook = load_workbook(excelFile)
    sheetObj = workBook.get_sheet_by_name(workBook.sheetnames[0])
    sheetObj.append((contentId,contentValue,str(contentId)+".png"))
    workBook.save(excelFile)
    workBook.close()
    return {"message":"add content success","contentId":contentId,"status":1}
#查询内容
def queryContent(contentId=None):
    resultList = []
    workBook = load_workbook(excelFile)
    sheetObj = workBook.get_sheet_by_name(workBook.sheetnames[0])
    maxRow = sheetObj.max_row
    if maxRow > 1:
        for i in range(2,maxRow+1):
            contentInfo = collections.OrderedDict()
            contentInfo["contentId"] = sheetObj.cell(row=i,column=1).value
            contentInfo["contentValue"] = sheetObj.cell(row=i,column=2).value
            contentInfo["contentImageFile"] = sheetObj.cell(row=i,column=3).value
            if contentId != None:
                if int(contentId) == int(sheetObj.cell(row=i,column=1).value):
                    resultList.append(contentInfo)
            else:
                resultList.append(contentInfo)
        return json.dumps(resultList,ensure_ascii=False,indent=4)
    else:
        return  resultList

def getRowAndFile(contentId):
    fileName = ""
    row = 0
    workBook = load_workbook(excelFile)
    sheetObj = workBook.get_sheet_by_name(workBook.sheetnames[0])
    maxRow = sheetObj.max_row
    if maxRow >1:
        for i in range(2,maxRow+1):
            if int(contentId) == int(sheetObj.cell(row=i,column=1).value):
                fileName = sheetObj.cell(row=i,column=3).value
                row = i
                break
    workBook.close()
    return row,fileName

#删除内容
def deleteContent(contentId):
    #打开Excel
    pythoncom.CoInitialize()#防止多线程执行出现文件并发操作，这里使用单线程操作
    workBook = win32com.client.Dispatch("Excel.Application").Workbooks.Open(excelFile)
    try:
        sheetObj = workBook.Worksheets("Sheet")
        row,fileName = getRowAndFile(contentId) #获取删除的行和文件
        if row > 1:
            #先删除数据
            sheetObj.Rows(row).Delete()
            #删除文件
            os.remove(imageRootPath+"\\"+fileName)
            return {"message":"delete success","status":1}
        else:
            return 0
        return {"message":"delete failed","status":1}
    except Exception,e:
        print str(e)
    finally:
        workBook.Close(SaveChanges=1)




















