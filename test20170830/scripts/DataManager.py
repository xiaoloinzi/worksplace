# encoding=utf-8
from openpyxl import load_workbook
import collections
#获取配置信息
def getConfigInfo(testCaseFile=None,index=None):
    configInfoDic = {}
    # 打开文件对象
    workBook = load_workbook(testCaseFile)
    # 获取sheet对象
    if index == None:
        index = 0
    sheet = workBook.get_sheet_by_name(workBook.sheetnames[index])
    configInfoDic["protocolType"] = sheet.cell(row=2, column=1).value
    configInfoDic["ip"] = sheet.cell(row=2, column=2).value
    configInfoDic["port"] = sheet.cell(row=2, column=3).value
    workBook.close()
    return configInfoDic
#从excel表格中获取接口信息
def getInterfaceInfo(testCaseFile=None,index=None):
    interfaceInfoDic = {}
    # 打开文件对象
    workBook = load_workbook(testCaseFile)
    # 获取sheet对象
    if index == None:
        index = 1
    sheet = workBook.get_sheet_by_name(workBook.sheetnames[index])
    row = sheet.max_row
    for i in range(2, row + 1):
        subInterfaceInfoDic = {}
        interfaceInfoDic["interfaceName"] = sheet.cell(row=i, column=1).value
        subInterfaceInfoDic["path"] = sheet.cell(row=i, column=2).value
        subInterfaceInfoDic["interfaceType"] = sheet.cell(row=i, column=3).value
        subInterfaceInfoDic["interfaceHeader"] = sheet.cell(row=i, column=4).value
        interfaceInfoDic[str(sheet.cell(row=i, column=1).value)] = subInterfaceInfoDic
    workBook.close()
    return interfaceInfoDic
#从excel表格中读取用例数据，并返回查询结果
def readTestCase(testCaseFile=None,index=None):
    # 打开文件对象
    workBook = load_workbook(testCaseFile)
    # 获取sheet对象
    if index == None:
        index = 2
    sheet =  workBook.get_sheet_by_name(workBook.sheetnames[index])
    # 获取最大行数
    row = sheet.max_row
    testCaseList = []
    for i in range(index, row + 1):
        testCaseDic = collections.OrderedDict()
        testCaseDic["testCaseId"] = sheet.cell(row=i, column=1).value
        testCaseDic["preContdtion"] = sheet.cell(row=i, column=2).value
        testCaseDic["preContdtionOutput"] = sheet.cell(row=i, column=3).value
        testCaseDic["testStep"] = sheet.cell(row=i, column=4).value
        testCaseDic["testStepOutput"] = sheet.cell(row=i, column=5).value
        testCaseDic["endCondition"] = sheet.cell(row=i, column=6).value
        testCaseDic["espectResult"] = sheet.cell(row=i, column=7).value
        testCaseDic["actualResult"] = sheet.cell(row=i, column=8).value
        testCaseDic["isSuccess"] = sheet.cell(row=i, column=9).value
        testCaseDic["failedReason"] = sheet.cell(row=i, column=10).value
        testCaseDic["remark"] = sheet.cell(row=i, column=11).value
        testCaseList.append(testCaseDic)
    workBook.close()
    return testCaseList